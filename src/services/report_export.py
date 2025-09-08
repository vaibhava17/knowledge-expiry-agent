"""
Report Export Service
Handles exporting reports to different formats (Excel, JSON, CSV)
"""

import json
import csv
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from loguru import logger

class ReportExporter:
    """Service for exporting reports in various formats"""
    
    def __init__(self):
        self.colors = {
            "header": "4472C4",
            "critical": "C5504B", 
            "high": "FF6B35",
            "medium": "FFB347",
            "low": "90EE90",
            "accent": "E7E6E6"
        }
    
    async def export_to_excel(
        self,
        report_data: Dict[str, Any],
        output_file: str,
        report_type: str = "comprehensive"
    ) -> bool:
        """
        Export report to Excel format with multiple sheets and formatting
        
        Args:
            report_data: Complete report data
            output_file: Output Excel file path
            report_type: Type of report (comprehensive, executive, detailed)
            
        Returns:
            Success status
        """
        try:
            logger.info(f"Exporting report to Excel: {output_file}")
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create sheets based on report type
            if report_type == "executive":
                self._create_executive_sheets(wb, report_data)
            elif report_type == "detailed":
                self._create_detailed_sheets(wb, report_data)
            else:  # comprehensive
                self._create_comprehensive_sheets(wb, report_data)
            
            # Save workbook
            wb.save(output_file)
            logger.info(f"Excel report exported successfully to {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False
    
    def _create_executive_sheets(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create executive summary sheets"""
        
        # Executive Summary sheet
        ws_summary = wb.create_sheet("Executive Summary")
        self._create_executive_summary_sheet(ws_summary, report_data)
        
        # Critical Findings sheet
        ws_findings = wb.create_sheet("Critical Findings")
        self._create_critical_findings_sheet(ws_findings, report_data)
        
        # Action Items sheet
        ws_actions = wb.create_sheet("Action Items")
        self._create_action_items_sheet(ws_actions, report_data)
    
    def _create_detailed_sheets(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create detailed analysis sheets"""
        
        # All Critical Points sheet
        ws_points = wb.create_sheet("All Critical Points")
        self._create_critical_points_sheet(ws_points, report_data)
        
        # Document Analysis sheet
        ws_docs = wb.create_sheet("Document Analysis")
        self._create_document_analysis_sheet(ws_docs, report_data)
        
        # Timeline Analysis sheet
        ws_timeline = wb.create_sheet("Timeline Analysis")
        self._create_timeline_sheet(ws_timeline, report_data)
    
    def _create_comprehensive_sheets(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create comprehensive report with all sheets"""
        
        # Executive sheets
        self._create_executive_sheets(wb, report_data)
        
        # Detailed sheets
        self._create_detailed_sheets(wb, report_data)
        
        # Additional analysis sheets
        ws_expiry = wb.create_sheet("Expiry Analysis")
        self._create_expiry_analysis_sheet(ws_expiry, report_data)
        
        ws_stats = wb.create_sheet("Statistics")
        self._create_statistics_sheet(ws_stats, report_data)
    
    def _create_executive_summary_sheet(self, ws, report_data: Dict[str, Any]):
        """Create executive summary sheet"""
        
        # Title
        ws['A1'] = "Knowledge Expiry Report - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill("solid", fgColor=self.colors["header"])
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        metadata = report_data.get("metadata", {})
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = metadata.get("generated_at", "Unknown")
        row += 1
        ws[f'A{row}'] = "Analysis Model:"
        ws[f'B{row}'] = metadata.get("analysis_model", "Unknown")
        row += 2
        
        # Key Metrics
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        key_metrics = report_data.get("executive_summary", {}).get("key_metrics", {})
        for metric, value in key_metrics.items():
            ws[f'A{row}'] = metric.replace("_", " ").title()
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Executive Summary Text
        ws[f'A{row}'] = "Executive Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        summary_text = report_data.get("executive_summary", {}).get("overview", "No summary available")
        ws[f'A{row}'] = summary_text
        ws.merge_cells(f'A{row}:D{row+5}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Apply formatting
        self._format_sheet(ws)
    
    def _create_critical_findings_sheet(self, ws, report_data: Dict[str, Any]):
        """Create critical findings sheet"""
        
        # Headers
        headers = ["Finding", "Impact", "Recommendation", "Priority"]
        self._create_header_row(ws, headers, 1)
        
        # Data
        row = 2
        critical_findings = report_data.get("critical_findings", [])
        
        for finding in critical_findings:
            ws[f'A{row}'] = finding.get("finding", "")
            ws[f'B{row}'] = finding.get("impact", "")
            ws[f'C{row}'] = finding.get("recommendation", "")
            ws[f'D{row}'] = finding.get("priority", "Medium")
            
            # Color code by priority
            priority = finding.get("priority", "Medium").lower()
            if priority in ["critical", "high"]:
                fill_color = self.colors.get("critical" if priority == "critical" else "high")
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill("solid", fgColor=fill_color)
            
            row += 1
        
        self._format_sheet(ws)
    
    def _create_action_items_sheet(self, ws, report_data: Dict[str, Any]):
        """Create action items sheet"""
        
        # Headers
        headers = ["Task", "Priority", "Owner", "Timeline", "Status"]
        self._create_header_row(ws, headers, 1)
        
        # Data
        row = 2
        action_items = report_data.get("recommendations", {}).get("action_items", [])
        
        for item in action_items:
            ws[f'A{row}'] = item.get("task", "")
            ws[f'B{row}'] = item.get("priority", "Medium")
            ws[f'C{row}'] = item.get("owner", "TBD")
            ws[f'D{row}'] = item.get("timeline", "TBD")
            ws[f'E{row}'] = "Pending"
            
            # Color code by priority
            priority = item.get("priority", "Medium").lower()
            if priority in ["critical", "high"]:
                fill_color = self.colors.get("critical" if priority == "critical" else "high")
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = PatternFill("solid", fgColor=fill_color)
            
            row += 1
        
        self._format_sheet(ws)
    
    def _create_critical_points_sheet(self, ws, report_data: Dict[str, Any]):
        """Create detailed critical points sheet"""
        
        # Headers
        headers = ["Description", "Category", "Urgency", "Document", "Confidence", "Context"]
        self._create_header_row(ws, headers, 1)
        
        # Data
        row = 2
        critical_points = report_data.get("critical_points", {}).get("detailed_list", [])
        
        for point in critical_points:
            ws[f'A{row}'] = point.get("description", "")
            ws[f'B{row}'] = point.get("category", "").title()
            ws[f'C{row}'] = point.get("urgency", "").title()
            ws[f'D{row}'] = point.get("document_filename", "")
            ws[f'E{row}'] = point.get("confidence_score", 0)
            ws[f'F{row}'] = point.get("context_snippet", "")[:100] + "..." if point.get("context_snippet", "") else ""
            
            # Color code by urgency
            urgency = point.get("urgency", "medium").lower()
            if urgency in self.colors:
                fill_color = self.colors[urgency]
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].fill = PatternFill("solid", fgColor=fill_color)
            
            row += 1
        
        self._format_sheet(ws)
    
    def _create_document_analysis_sheet(self, ws, report_data: Dict[str, Any]):
        """Create document analysis sheet"""
        
        # File type distribution
        ws['A1'] = "Document Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        ws[f'A{row}'] = "File Type Distribution"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        doc_analysis = report_data.get("document_analysis", {})
        file_types = doc_analysis.get("file_type_distribution", {})
        
        ws[f'A{row}'] = "File Type"
        ws[f'B{row}'] = "Count"
        self._create_header_row(ws, ["File Type", "Count"], row)
        row += 1
        
        for file_type, count in file_types.items():
            ws[f'A{row}'] = file_type.upper()
            ws[f'B{row}'] = count
            row += 1
        
        row += 2
        
        # Confidence distribution
        ws[f'A{row}'] = "Confidence Score Distribution"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        conf_dist = doc_analysis.get("confidence_distribution", {})
        ws[f'A{row}'] = "Confidence Level"
        ws[f'B{row}'] = "Count"
        self._create_header_row(ws, ["Confidence Level", "Count"], row)
        row += 1
        
        for level, count in conf_dist.items():
            ws[f'A{row}'] = level
            ws[f'B{row}'] = count
            row += 1
        
        self._format_sheet(ws)
    
    def _create_timeline_sheet(self, ws, report_data: Dict[str, Any]):
        """Create timeline analysis sheet"""
        
        ws['A1'] = "Timeline Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        timeline = report_data.get("timeline_analysis", {})
        timeline_cats = timeline.get("timeline_categories", {})
        
        ws[f'A{row}'] = "Timeline Category"
        ws[f'B{row}'] = "Items Count"
        self._create_header_row(ws, ["Timeline Category", "Items Count"], row)
        row += 1
        
        for category, count in timeline_cats.items():
            ws[f'A{row}'] = category.replace("_", " ").title()
            ws[f'B{row}'] = count
            
            # Color code by urgency
            if "immediate" in category:
                fill_color = self.colors["critical"]
            elif "30" in category:
                fill_color = self.colors["high"]
            elif "90" in category:
                fill_color = self.colors["medium"]
            else:
                fill_color = self.colors["low"]
            
            ws[f'A{row}'].fill = PatternFill("solid", fgColor=fill_color)
            ws[f'B{row}'].fill = PatternFill("solid", fgColor=fill_color)
            
            row += 1
        
        self._format_sheet(ws)
    
    def _create_expiry_analysis_sheet(self, ws, report_data: Dict[str, Any]):
        """Create expiry indicators analysis sheet"""
        
        ws['A1'] = "Knowledge Expiry Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        expiry_analysis = report_data.get("expiry_analysis", {})
        
        # Summary stats
        ws[f'A{row}'] = "Total Points with Expiry Indicators:"
        ws[f'B{row}'] = expiry_analysis.get("total_points_with_indicators", 0)
        row += 2
        
        # Most common indicators
        ws[f'A{row}'] = "Most Common Expiry Indicators"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Indicator"
        ws[f'B{row}'] = "Frequency"
        self._create_header_row(ws, ["Indicator", "Frequency"], row)
        row += 1
        
        common_indicators = expiry_analysis.get("most_common_indicators", [])
        for indicator, frequency in common_indicators[:10]:
            ws[f'A{row}'] = indicator
            ws[f'B{row}'] = frequency
            row += 1
        
        self._format_sheet(ws)
    
    def _create_statistics_sheet(self, ws, report_data: Dict[str, Any]):
        """Create statistics summary sheet"""
        
        ws['A1'] = "Database Statistics"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Document statistics
        doc_stats = report_data.get("appendix", {}).get("database_statistics", {})
        ws[f'A{row}'] = "Document Statistics"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        for key, value in doc_stats.items():
            ws[f'A{row}'] = key.replace("_", " ").title()
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Vector DB statistics
        vector_stats = report_data.get("appendix", {}).get("vector_db_statistics", {})
        ws[f'A{row}'] = "Vector Database Statistics"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        for key, value in vector_stats.items():
            ws[f'A{row}'] = key.replace("_", " ").title()
            ws[f'B{row}'] = value
            row += 1
        
        self._format_sheet(ws)
    
    def _create_header_row(self, ws, headers: List[str], row: int):
        """Create formatted header row"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=self.colors["header"])
            cell.alignment = Alignment(horizontal="center")
    
    def _format_sheet(self, ws):
        """Apply general formatting to sheet"""
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    async def export_to_json(self, report_data: Dict[str, Any], output_file: str) -> bool:
        """Export report to JSON format"""
        try:
            logger.info(f"Exporting report to JSON: {output_file}")
            
            # Ensure output directory exists
            Path(output_file).parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"JSON report exported successfully to {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            return False
    
    async def export_to_csv(self, report_data: Dict[str, Any], output_file: str) -> bool:
        """Export critical points to CSV format"""
        try:
            logger.info(f"Exporting report to CSV: {output_file}")
            
            # Ensure output directory exists
            Path(output_file).parent.mkdir(parents=True, exist_ok=True)
            
            # Export critical points to CSV
            critical_points = report_data.get("critical_points", {}).get("detailed_list", [])
            
            if not critical_points:
                logger.warning("No critical points to export to CSV")
                return False
            
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                fieldnames = [
                    'description', 'category', 'urgency', 'document_filename',
                    'confidence_score', 'context_snippet', 'last_updated_date'
                ]
                
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for point in critical_points:
                    # Filter and clean the data for CSV
                    csv_row = {
                        'description': point.get('description', ''),
                        'category': point.get('category', ''),
                        'urgency': point.get('urgency', ''),
                        'document_filename': point.get('document_filename', ''),
                        'confidence_score': point.get('confidence_score', 0),
                        'context_snippet': point.get('context_snippet', '')[:200] + '...' if point.get('context_snippet', '') else '',
                        'last_updated_date': point.get('last_updated_date', '')
                    }
                    writer.writerow(csv_row)
            
            logger.info(f"CSV report exported successfully to {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return False